#!/usr/bin/env python
# -*- coding: utf-8 -*-

import copy
import os
import re
from enum import Enum
import datetime
import openpyxl
from openpyxl.worksheet import *


class Severity(Enum):
    """The values of column Severity"""
    Critical = (0, 'Critical')
    Major = (1, 'Major')
    Minor = (2, 'Minor')
    Warning = (3, 'Warning')
    Ok = (10, 'Ok')

    def __str__(self):
        return self.value[1]


class Check(Enum):
    """Header of checks"""
    Caption = 0
    Command = 1
    Output = 2
    Element = 3
    AlarmsReference = 4


class Alarm(Enum):
    """Header of alarms"""
    specificProblem = 0
    eventType = 1
    probableCause = 2
    perceivedSeverity = 3
    managedObject = 4
    RNCNodeType = 5


class ZbCheckRow():
    """ строка страницы файла Support Report """
    def __init__(self, checkname, order, severity=Severity.Ok, observation='No alarms', dateof='', nodename=''):
        super(ZbCheckRow, self).__init__()
        self.CheckName = checkname
        self.Severity = severity
        self.Observation = observation
        self.Order = order
        self.DateOf = dateof
        self.NodeName = nodename
        self.alarmsDetail = []
        self.alarmsCritical = 0
        self.alarmsMajor = 0
        self.alarmsMinor = 0
        self.alarmsWarning = 0
        self.alarmsTotal = 0
        self.alarmsCollision = 0

    def __str__(self):
        return '\t'.join([str(self.Order), self.CheckName, str(self.Severity), self.Observation])


def copy_rows(self, row_idx, cnt, above=False, copy_style=True, fill_formulae=True):
    """Inserts new (empty) rows into worksheet at specified row index.
    :param self: itself
    :param row_idx: Row index specifying where to insert new rows.
    :param cnt: Number of rows to insert.
    :param above: Set True to insert rows above specified row index.
    :param copy_style: Set True if new rows should copy style of immediately above row.
    :param fill_formulae: Set True if new rows should take on formula from immediately above row, filled with references new to rows.
    Usage:
    * insert_rows(2, 10, above=True, copy_style=False)
    """
    CELL_RE = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")
    row_idx = row_idx - 1 if above else row_idx

    def replace(m):
        row = m.group('row')
        prefix = "$" if row.find("$") != -1 else ""
        row = int(row.replace("$",""))
        row += cnt if row > row_idx else 0
        return m.group('col') + prefix + str(row)
    # First, we shift all cells down cnt rows...
    old_cells = set()
    old_fas = set()
    new_cells = dict()
    new_fas = dict()
    for c in self._cells.values():
        old_coor = c.coordinate
        # Shift all references to anything below row_idx
        if c.data_type == Cell.TYPE_FORMULA:
            c.value = CELL_RE.sub(
                replace,
                c.value
            )
            # Here, we need to properly update the formula references to reflect new row indices
            if old_coor in self.formula_attributes and 'ref' in self.formula_attributes[old_coor]:
                self.formula_attributes[old_coor]['ref'] = CELL_RE.sub(
                    replace,
                    self.formula_attributes[old_coor]['ref']
                )
        # Do the magic to set up our actual shift    
        if c.row > row_idx:
            old_coor = c.coordinate
            old_cells.add((c.row,c.col_idx))
            c.row += cnt
            new_cells[(c.row,c.col_idx)] = c
            if old_coor in self.formula_attributes:
                old_fas.add(old_coor)
                fa = self.formula_attributes[old_coor].copy()
                new_fas[c.coordinate] = fa
    for coor in old_cells:
        del self._cells[coor]
    self._cells.update(new_cells)
    for fa in old_fas:
        del self.formula_attributes[fa]
    self.formula_attributes.update(new_fas)
    # Next, we need to shift all the Row Dimensions below our new rows down by cnt...
    for row in range(len(self.row_dimensions)-1+cnt,row_idx+cnt,-1):
        new_rd = copy.copy(self.row_dimensions[row-cnt])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        del self.row_dimensions[row-cnt]
    # Now, create our new rows, with all the pretty cells
    row_idx += 1
    for row in range(row_idx,row_idx+cnt):
        # Create a Row Dimension for our new row
        new_rd = copy.copy(self.row_dimensions[row-1])
        new_rd.index = row
        self.row_dimensions[row] = new_rd
        for col in range(1,self.max_column):
            col = get_column_letter(col)
            cell = self.cell('%s%d'%(col,row))
            cell.value = self.cell('%s%d'%(col,row-1)).value
            source = self.cell('%s%d'%(col,row-1))
            if copy_style:
                cell.number_format = source.number_format
                cell.font        = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border    = source.border.copy()
                cell.fill        = source.fill.copy()
            if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
                s_coor = source.coordinate
                if s_coor in self.formula_attributes and 'ref' not in self.formula_attributes[s_coor]:
                    fa = self.formula_attributes[s_coor].copy()
                    self.formula_attributes[cell.coordinate] = fa
                # print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
                cell.value = re.sub(
                    "(\$?[A-Z]{1,3}\$?)%d"%(row - 1),
                    lambda m: m.group(1) + str(row),
                    source.value
                )     
                cell.data_type = Cell.TYPE_FORMULA
    # Check for Merged Cell Ranges that need to be expanded to contain new cells
    for cr_idx, cr in enumerate(self.merged_cell_ranges):
        self.merged_cell_ranges[cr_idx] = CELL_RE.sub(
            replace,
            cr
        )
Worksheet.copy_rows = copy_rows


class ZbAnalyser():
    """zbAnalyser! И этим всё сказано"""
    def __init__(self):
        super(ZbAnalyser, self).__init__()
        self.currentTemplate = '160123_t.xlsx'
        self.referenceError = 'Alarms_and_events.xlsx'
        self.dirs = { 'inputDir': './input', 'outputDir': './output', 'logDir': './log' }
        self.checks = (('Check active Alarms', 'alt', 
                        r'(?si)Date & Time \(Local\) +S +Specific Problem +MO \(Cause/AdditionalInfo\)\n'
                        r'={10,}\n'
                        r'(.*?)\n?'
                        r'>>> Total: \d+ Alarms \(\d+ Critical, \d+ Major\)',
                        r'20\d{2}-\d{2}-\d{2} \d{2}:\d{2}:\d{2} (\w) ((?:\w+ ?)+) +(.*)', 'Alarms_and_events.xlsx'),
                       ('Check Event and System Logs', 'lgesmr 7d',
                        r'(?si)Timestamp \(UTC\) +Type +Merged Log Entry\n'
                        r'={10,}\n'
                        r'(.*)', r'(?i)[\d-]+ [\d:]+ +\w+ +(?:(?:(?:\w+=[\w-]+),)+(\w+=[\w-]+)|(?:Crash on (\d+), '
                                 r'device=(\d+) \w+)) +(.+)', ''),
                       ('Check Node Restart and System Downtime', 'lgd',
                        r'(?si)Timestamp \(UTC\) +RestartType/Reason +Configuration Version +SwRelease +CPP Downtime +'
                        r'Appl. Downtime +JVM Downtime\n'
                        r'={10,}\n'
                        r'(.+)\n+'
                        r'Node uptime since last restart: \d+ \w+ \((?:(\d+) days)?,? ?(?:(\d+) hours)?',
                        r'(?i)(\d{4})-(\d{2})-(\d{2}) [\d:]+ Spontaneous', ''),
                       ('Check Date and Time Synchronization', 'lh coremp readclock',
                        r'(?si)\d{6}-\d{2}:\d{2}:\d{2} [\w \d./=]+\n(.+)',
                        r'\$ lhsh 00\d{2}00 readclock\n\d+: Date: 20(\d{2})-(\d{2})-(\d{2})', ''),
                       ('Check Network Synchronization', ('get Synchronization=1', 'st tusync'), '(.*)', '', ''),
                       ('Check the M3UA Associations', 'st m3ua',
                        '(?si)Proxy +Adm State +Op. +State +MO\n={10,}\n'
                        '(.*?)={10,}\n'
                        'Total: \d+ MOs',
                        r'(?i) +\d+ +\d+ \(DISABLED\) +(?:[\w\d]+=[\w\d]+,)*M3uAssociation=(\w{2})[\d\w]+', ''),
                       ('Check RNC CC, DC and PDR devices', 'std',
                        '(?i)-{10,}\nType +%Up +Total +Enabled\(1\) +Disabled\(0\) +Locked\(L\) +Active\(A\) +'
                        'Idle\(I\) +Busy\(B\) +Unallocated\n-{10,}\n'
                        '((?:\w+ +\d+% +\d+ +\d+ +\d+ +\d+ +\d+ +\d+ +\d+ +\d+\n)+)-{10,}\n'
                        'TOT +\d+% +\d+ +\d+ +\d+ +\d+ +\d+ +\d+ +\d+ +\d',
                        '(?i)(\w+) +(\d+)% +(\d+) +(\d+) +(\d+) +(\d+) +(\d+) +(\d+) +(\d+) +(\d+)', ''),
                       ('Check IubLink and Utrancell resource Status', 'strt',
                        r'(?si)Following \d+ sites are up:.*'
                        r'Following \d+ sites are totally or partially unavailable:.*-{10,}\n+'
                        r'(.*)',
                        r'([\w ]+): +(\d+) of +(\d+) [\w ]+\(([\d.]+) %\)', ''),
                       ('Check RANAP and Iu link', 'st ranap',
                        '(?si)Proxy +Adm +State +Op. State +MO\n'
                        '={10,}\n'
                        '(.*?)={10,}\n'
                        'Total: \d+ MOs', '(?si) *\d+ +[\d\w]+ \(DISABLED\) +((?:[\w\d_]+=[\w\d_]+,?)+)', ''),
                       ("Check CV's stored on RNC", 'cvls',
                        r"(?i)>>> Total: (\d+ CV's, \d+ UP's)", "(?i)(\d+)[\w', ]+(\d+)", ''),
                       ('Check CV Database inconsistency', 'dbc', '(?si)(Conclusion: the database is (?:NOT )?OK)',
                        '(?si)(Conclusion: the database is NOT OK)', ''),
                       ('Current software level', 'cvcu',
                        '(?i)((?:[\w+ ]+: +[\w%=/]+ +[\w/]+ +W[\d.]+ \([\w\d.-]+\)\n|-{10,}\n)+)',
                        '[\w+ ]+: +[\w%=/]+ +[\w/]+ +W([\d.]+) \([\w\d.-]+\)', ''),
                       ('Check ethernet connectivity (i.e. Internal_IP_Transport Vlan)', ('steg', 'stip'), 'EXCEPTION',
                        'EXCEPTION', ''),
                       ('Check CC/DC/PDR allocation', ('lh cenmp drhcendh cc', 'lh cenmp drhcendh dc',
                                                       'lh cenmp drhcendh pdr'),
                        'EXCEPTION', 'EXCEPTION', ''),
                       ('Check for disable Mos', 'st all 1\.\*0', '(?is)Proxy +Adm +State +Op. +State +MO\n={10,}\n'
                                                                  '(.*?)\n?={10,}\nTotal: \d+ MOs', '(.+)', ''),
                       ('Health check result', 'get ManagedElement=1 healthCheckResult\|healthCheckSchedule',
                        r'(?si)={10,}\n'
                        r'MO +Attribute +Value\n={10,}\n'
                        r'(.*?)\n?'
                        r'={10,}\n'
                        r'Total: \d+ Mos',
                        r'(?i)ManagedElement=\d+ +healthCheckResult Struct\{\d\}.*\n?'
                        r'(?: >>> 1.healthCheckResultCode = (\d+ \(\w+\)).*)?\n?'
                        r'(?: >>> 2.message = (.*))?\n?'
                        r'(?: >>> 3.startTime = [\d-]+ [\d:]+)?', ''),
                       ('Health check scheduler', 'get ManagedElement=1 healthCheckResult\|healthCheckSchedule',
                        r'(?si)={10,}\n'
                        r'MO +Attribute +Value\n={10,}\n'
                        r'(.*?)\n?'
                        r'={10,}\n'
                        r'Total: \d+ Mos',
                        r'ManagedElement=\d+ +healthCheckSchedule t\[(\d+)\].*\n?'
                        r'(?: >>> Struct\[\d\] +has \d+.*)?\n?'
                        r'(?: >>> 1[.]time = \d{2}:\d{2})?\n?'
                        r'(?: >>> 2[.]weekday = \d+ \(\w+\))?', ''),
                       ('Check repartition of IubLinks and Cells', 'lkra',
                        '(?is)Sr +Mod +S +GPB +nIub +CellGPB +CellCC +nCC\n-{10,}\n'
                        '(.*?)\n?-{10,}\n+Cell repartition by Board:', '', ''))
        self.output = []
        self.wb = None
        self.log = None
        self.alarms = None
        self.alarmsReferenceName = ''
        self.logdate = None
        if not os.path.exists(self.dirs['inputDir']):
            os.mkdir(self.dirs['inputDir'])
        if not os.path.exists(self.dirs['outputDir']):
            os.mkdir(self.dirs['outputDir'])
        if not os.path.exists(self.dirs['logDir']):
            os.mkdir(self.dirs['logDir'])

    def init_alarms(self):
        wb = openpyxl.load_workbook(filename=self.referenceError)
        ws = wb['Alarms']
        self.alarms = []
        headerfounded = False
        for row in ws.iter_rows():
            if headerfounded:
                b = []
                for cell in row:
                    if cell.column in ('A', 'B', 'C', 'D', 'E', 'F'):
                        b.append(cell.value.strip(' '))
                if b[0] is not None:
                    self.alarms.append(tuple(b))
            if row[0].value and row[0].value.lower().strip(' ') == 'specificproblem':
                headerfounded = True
        self.alarms = tuple(self.alarms)
        return self.alarms

    def check14(self, nextStr, output):
        type = ''
        if output.find('drhcendh cc') >= 0:
            type = 'cc'
        elif output.find('drhcendh dc') >= 0:
            type = 'dc'
        elif output.find('drhcendh pdr') >= 0:
            type = 'pdr'
        n, m, p = 0, 0, 1.
        idisunique = True
        e = re.compile(r'(?is)(\d+): +deviceId +devFroId +boardPiuFroId +admState +opState +capability +subrack '
                       r'+servingRhModuleId +linkHandlerName +ptmLm +faultTable +state +msgBoard'
                       r'\n(.*?)\n\1: Summary of resource handlers:')
        if e.search(output):
            for outputlines in e.findall(output):
                e1 = re.compile(r'(?i).*?(\w+) (\[.*\])')
                if e1.search(outputlines[1]):
                    for status, msgboard in e1.findall(outputlines[1]):
                        if msgboard.lower().strip(' ') != '[linkestablished synced allocatedsp allocatedrh ]':
                            n += 1
                        if status.lower().strip(' ') == 'idle':
                            m += 1
                    p = m / len(e1.findall(outputlines[1]))
        e = re.compile(r'(?is)\d+: +moduleId +boardPiuFroId +moduleRole +connected +subrack +noSpDev +capability +'
                       r'allocatedShare +properShare +pendingRebalance\n'
                       r'(.*?)\n?={10,}')
        if e.search(output):
            mset = set()
            for outputlines in e.findall(output):
                e1 = re.compile(r'(?is)^\d+: +(\d+)')
                if e1.search(outputlines):
                    for moduleid in e1.findall(outputlines):
                        if moduleid in mset:
                            idisunique = False
                            break
                        mset.add(moduleid)
        if n > 2:
            nextStr.Severity = Severity.Critical
        elif n > 0 or idisunique is False:
            if nextStr.Severity.value[0] > Severity.Major.value[0]:
                nextStr.Severity = Severity.Major
        elif p >= 0.5:
            if nextStr.Severity.value[0] > Severity.Warning.value[0]:
                nextStr.Severity = Severity.Warning
        nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + '%s status are %s' % \
        (type.upper(), 'NOK' if nextStr.Severity != Severity.Ok else 'OK')
        return nextStr

    def parseLog(self, nodename):
        if self.log is None:
            print('No log!')
            return
        logdatere = re.search(r'Logging to file [/\w\d]+/(\d{4}-\d{2}-\d{2})', self.log)
        if logdatere:
            self.logdate = logdatere.group(1)
        for num, check in enumerate(self.checks):
            nextStr = ZbCheckRow(checkname=check[Check.Caption.value], order=num, nodename=nodename)
            nextStr.Observation = ''
            commandRegExp = r'(?m)^(?:[\w\d.]+)> %s\n((?:.*\n?(?!(?:^[\w\d.]+)>))*)'
            if (check[Check.AlarmsReference.value] != '' and
                os.path.exists(check[Check.AlarmsReference.value]) and
               self.alarmsReferenceName != check[Check.AlarmsReference.value]):
                self.alarmsReferenceName = check[Check.AlarmsReference.value]
                self.init_alarms()
            outputREO = re.compile(commandRegExp % (check[Check.Command.value] if not isinstance(check[Check.Command.value], tuple) else str(check[Check.Command.value]).replace('(','(?:', 1).replace(", ","|").replace("'","")))
            if outputREO.search(self.log) is None:
                print('%s - outputRE is fail!' % nextStr.CheckName)
                continue
            for output in outputREO.findall(self.log):
                if output is None:
                    print('%s - Command RegExp fail' % nextStr.CheckName)
                    continue
                commandDateRE = re.search(r'(\d{6})-\d{2}:\d{2}:\d{2}', output)
                if commandDateRE:
                    if nextStr.DateOf != '' and nextStr.DateOf != commandDateRE.group(1):
                        print("Command date is different!")
                    nextStr.DateOf = commandDateRE.group(1)
                if nextStr.CheckName == 'Check CC/DC/PDR allocation' :
                    nextStr = self.check14(nextStr, output)
                    continue
                if nextStr.CheckName == 'Check ethernet connectivity (i.e. Internal_IP_Transport Vlan)' :
                    novlans = False
                    prioequal = False
                    pbitqmapnequal = False
                    dscppbitmapnequal = False
                    edgeoff = None
                    header = re.compile(r'(?i)(Board +)(.*)(Speed +)(.*)(Prio +)(.*)(Edge +)(PbitQMap +)(.*)(Vlans)').search(output)
                    outputLinesRE = re.compile(r'(?si)Board +.*Speed.*Prio.*Edge +PbitQMap.*Vlans\n={10,}\n'
                                               r'(.*?)={10,}\nTotal: \d+ MOs')
                    if outputLinesRE.search(output):
                        iboard = 0
                        ispeed = 1
                        iprio = 2
                        iedge=3
                        ipbitqmap = 4
                        ivlans = 5
                        pbitqmap = None
                        prioequal = True
                        prio = None
                        for outputLines in outputLinesRE.findall(output):
                            elementRE = re.compile(r'(?P<board>.{%d}).{%d}(?P<speed>.{%d}).{%d}(?P<prio>.{%d}).{%d}'
                                                   r'(?P<edge>.{%d})(?P<pbitqmap>.{%d}).{%d}(?P<vlans>.*)' %
                                                   (len(header.group(1)), len(header.group(2)), len(header.group(3)),
                                                    len(header.group(4)), len(header.group(5)), len(header.group(6)),
                                                    len(header.group(7)), len(header.group(8)), len(header.group(9))))
                            for element in elementRE.findall(outputLines):
                                if element[iedge].strip(' ').lower() == 'edge_on':
                                    edgeoff = False
                                elif edgeoff is None:
                                    edgeoff = True
                                if pbitqmap is None:
                                    pbitqmap = element[ipbitqmap].strip(' ').lower()
                                elif pbitqmap != element[ipbitqmap].strip(' ').lower():
                                    pbitqmapnequal = True
                                if prio is None:
                                    prio = element[iprio].strip(' ').lower()
                                elif prio != element[iprio].strip(' ').lower():
                                    prioequal = False
                                if ((element[iboard].strip(' ').lower() == 'cmxb' and
                                     element[ispeed].strip(' ').lower() == 'nolink' and
                                     element[ivlans].strip(' ').lower() == '') or
                                    (element[iboard].strip(' ').lower() == 'ipg' and
                                     element[ivlans].strip(' ').lower() == '')):
                                    novlans = True
                    header = re.compile(r'(?i)(Board +)(.*)(Speed +)(.*)(Vlans +)(DscpPbitMap)').search(output)
                    outputLinesRE = re.compile(r'(?si)Board.*Speed.*Vlans +DscpPbitMap\n={10,}\n'
                                               r'(.*?)={10,}\nTotal: \d+ MOs')
                    if outputLinesRE.search(output):
                        iboard = 0
                        ispeed = 1
                        ivlans = 2
                        idscppbitmap = 3
                        dscppbitmap = None
                        for outputLines in outputLinesRE.findall(output):
                            elementRE = re.compile(r'(?P<board>.{%d}).{%d}(?P<speed>.{%d}).{%d}(?P<vlans>.{%d}).{%d}'
                                                   r'(?P<dscppbitmap>.*)' %
                                                   (len(header.group(1)), len(header.group(2)), len(header.group(3)),
                                                    len(header.group(4)), len(header.group(5)), len(header.group(6))))
                            for element in elementRE.findall(outputLines):
                                if dscppbitmap is None:
                                    dscppbitmap = element[idscppbitmap].strip(' ').lower()
                                elif dscppbitmap != element[idscppbitmap].strip(' ').lower():
                                    dscppbitmapnequal = True
                                if ((element[iboard].strip(' ').lower() == 'cmxb' and
                                     element[ispeed].strip(' ').lower() == 'nolink' and
                                     element[ivlans].strip(' ').lower() == '') or
                                    (element[iboard].strip(' ').lower() == 'ipg' and
                                     element[ivlans].strip(' ').lower() == '')):
                                    novlans = True
                    if novlans:
                        nextStr.Severity = Severity.Critical
                    elif prioequal:
                        if nextStr.Severity.value[0] > Severity.Major.value[0]:
                            nextStr.Severity = Severity.Major
                    elif pbitqmapnequal or dscppbitmapnequal:
                        if nextStr.Severity.value[0] > Severity.Minor.value[0]:
                            nextStr.Severity = Severity.Minor
                    elif edgeoff:
                        if nextStr.Severity.value[0] > Severity.Warning.value[0]:
                            nextStr.Severity = Severity.Warning
                    if nextStr.Severity != Severity.Ok:
                        nextStr.Observation = 'NOk'
                    else:
                        nextStr.Observation += 'Ok'
                    continue
                outputLinesRE = re.search(check[Check.Output.value], output)
                if outputLinesRE is None:
                    print('%s - outputLinesRE is fail!' % nextStr.CheckName)
                    continue
                outputLines = outputLinesRE.group(1)
                elementRE = re.compile(check[Check.Element.value])
                if check[Check.Command.value] in [self.checks[0][Check.Command.value]]:
                    if elementRE.search(outputLines):
                        for element in elementRE.findall(outputLines):
                            if self.alarmsReferenceName == check[Check.AlarmsReference.value] and self.alarms is not None:
                                for alarm in self.alarms:
                                    if element[1].lower().strip(' ') == alarm[Alarm.specificProblem.value].lower():
                                        if element[0] == 'c' and alarm[Alarm.perceivedSeverity.value].lower() == 'critical':
                                            nextStr.alarmsCritical += 1
                                            nextStr.alarmsDetail.append(element[1])
                                        elif element[0] == 'M' and alarm[Alarm.perceivedSeverity.value].lower() == 'major':
                                            nextStr.alarmsMajor += 1
                                            nextStr.alarmsDetail.append(element[1])
                                        elif element[0] == 'm' and alarm[Alarm.perceivedSeverity.value].lower() == 'minor':
                                            nextStr.alarmsMinor += 1
                                        elif element[0] == 'w' and alarm[Alarm.perceivedSeverity.value].lower() == 'warning':
                                            nextStr.alarmsWarning += 1
                                        else:
                                            nextStr.alarmsCollision += 1
                                            print('%s - Unknown perceivedSeverity!' % nextStr.CheckName)
                                        # print(element)
                                        break
                    nextStr.alarmsTotal = nextStr.alarmsCritical + nextStr.alarmsMajor + nextStr.alarmsMinor + \
                                          nextStr.alarmsWarning + nextStr.alarmsCollision
                    if nextStr.alarmsTotal > 0:
                        nextStr.Observation = ('\n' if nextStr.Observation != '' else '') + 'Total %d alarms:' % \
                                                                                            nextStr.alarmsTotal
                        comma = False
                        if nextStr.alarmsCritical > 0:
                            nextStr.Observation += ' %d critical' % nextStr.alarmsCritical
                            comma = True
                            nextStr.Severity = Severity.Critical
                        if nextStr.alarmsMajor > 0:
                            nextStr.Observation += (',' if comma else '') + ' %d major' % nextStr.alarmsMajor
                            comma = True
                            if nextStr.Severity.value[0] > Severity.Major.value[0]:
                                nextStr.Severity = Severity.Major
                        if nextStr.alarmsMinor > 0:
                            nextStr.Observation += (',' if comma else '') + ' %d minor' % nextStr.alarmsMinor
                            comma = True
                            if nextStr.Severity.value[0] > Severity.Minor.value[0]:
                                nextStr.Severity = Severity.Minor
                        if nextStr.alarmsWarning > 0:
                            nextStr.Observation += (',' if comma else '') + ' %d warning' % nextStr.alarmsWarning
                            comma = True
                            if nextStr.Severity.value[0] > Severity.Warning.value[0]:
                                nextStr.Severity = Severity.Warning
                if check[Check.Command.value] == self.checks[1][Check.Command.value]:
                    if elementRE.search(outputLines):
                        sum = 0
                        MOs = set()
                        for element in elementRE.findall(outputLines):
                            if element[3] is not None and element[3].lower().find('ranap_cninitiatedresetresource') >= 0:
                                nextStr.Severity = Severity.Critical
                                MOs.add(element[0])
                                sum += 1
                        if sum != 0:
                            if nextStr.Observation != '':
                                nextStr.Observation += '\n'
                            nextStr.Observation += 'Ranap_CNInitiatedResetResource %s sum: %d' % \
                                                   (str(MOs).strip('{}').replace("'",""), sum)
                        sum = 0
                        MOs = set()
                        for element in elementRE.findall(outputLines):
                            if element[3] is not None and element[3].lower().find('ipethpacketdatarouter_cnnotresponding' +
                                                                                  'togtpecho') >= 0:
                                if nextStr.Severity.value[0] > Severity.Major.value[0]:
                                    nextStr.Severity = Severity.Major
                                MOs.add(element[0])
                                sum += 1
                        if sum != 0:
                            if nextStr.Observation != '':
                                nextStr.Observation += '\n'
                            nextStr.Observation += 'IpEthPacketDataRouter_CnNotRespondingToGTPEcho %s sum: %d' % \
                                                   (str(MOs).strip('{}').replace("'",""), sum)
                        sum = 0
                        MOs = set()
                        prevdevice = ''
                        for element in elementRE.findall(outputLines):
                            if element[1] is None or element[1] == '':
                                continue
                            if prevdevice != '' and prevdevice != element[2]:
                                nextStr.Severity = Severity.Critical
                            elif int(element[1]) > 1 and nextStr.Severity.value[0] > Severity.Major.value[0]:
                                nextStr.Severity = Severity.Major
                            elif int(element[1]) == 1 and nextStr.Severity.value[0] > Severity.Minor.value[0]:
                                nextStr.Severity = Severity.Minor
                            MOs.add(', '.join(['Crash on %s' % element[1], 'device=%s' % element[2]]))
                            sum += 1
                        if sum != 0:
                            if nextStr.Observation != '':
                                nextStr.Observation += '\n'
                            nextStr.Observation += '%s sum: %d' % (str(MOs).strip('{}').replace("'",""), sum)
                        sum = 0
                        MOs = set()
                        for element in elementRE.findall(outputLines):
                            if element[3] is not None and element[3].lower().find('a non-local mau has been chosen as the' +
                                                                                  ' active client') >= 0:
                                if nextStr.Severity.value[0] > Severity.Warning.value[0]:
                                    nextStr.Severity = Severity.Warning
                                MOs.add(element[0])
                                sum += 1
                        if sum != 0:
                            if nextStr.Observation != '':
                                nextStr.Observation += '\n'
                            nextStr.Observation += 'A Non-Local MAU Has Been Chosen as the Active Client %s sum: %d' % \
                                                   (str(MOs).strip('{}').replace("'",""), sum)
                if check[Check.Command.value] in [self.checks[2][Check.Command.value]]:
                    sum = 0
                    if elementRE.search(outputLines):
                        for element in elementRE.findall(outputLines):
                            opdate = datetime.date(int(element[0]), int(element[1]), int(element[2]))
                            repdate = datetime.date(2000 + int(nextStr.DateOf[0:2]), int(nextStr.DateOf[2:4]), int(nextStr.DateOf[4:6]))
                            if (repdate - opdate).days <= 14:
                                sum += 1
                    if sum > 1:
                        nextStr.Severity = Severity.Critical
                    if sum == 1:
                        nextStr.Severity = Severity.Major
                    nextStr.Observation = 'Node uptime since last restart: %s days, %s hours' % \
                                          (outputLinesRE.group(2), outputLinesRE.group(3))
                if check[Check.Command.value] in [self.checks[3][Check.Command.value]]:
                    if elementRE.search(outputLines):
                        for element in elementRE.findall(outputLines):
                            if element[0]+element[1]+element[2] != nextStr.DateOf and nextStr.Severity.value[0] > Severity.Minor.value[0]:
                                nextStr.Severity = Severity.Minor
                                if nextStr.Observation != '':
                                    nextStr.Observation += '\n'
                                nextStr.Observation += 'Please check NTP'
                                break
                if check[Check.Command.value] in [self.checks[4][Check.Command.value]]:
                    pass
                if check[Check.Command.value] in [self.checks[5][Check.Command.value]]:
                    if elementRE.search(outputLines):
                        cs, ps, rs, MOs = 0, 0, 0, 0
                        for element in elementRE.findall(outputLines):
                            if element.lower() == 'cs':
                                cs += 1
                            elif element.lower() == 'ps':
                                ps += 1
                            elif element.lower() == 'rs':
                                rs += 1
                            MOs += 1
                        if cs >= 2 or ps >= 2:
                            nextStr.Severity = Severity.Critical
                        elif cs == 1 or ps == 1 or rs > 0 and nextStr.Severity.value[0] > Severity.Major.value[0]:
                            nextStr.Severity = Severity.Major
                        elif MOs > 0 and nextStr.Severity.value[0] > Severity.Minor.value[0]:
                            nextStr.Severity = Severity.Minor
                        if MOs > 0:
                            nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + 'Num of failed M3UA: %d' % MOs
                if check[Check.Command.value] in [self.checks[6][Check.Command.value]]:
                    if elementRE.search(outputLines):
                        disabled, unallocate, pdr, cc, dc = 0, 0, 0, 0, 0
                        for element in elementRE.findall(outputLines):
                            if element[0].lower() == 'pdr':
                                pdr = int(element[1])
                            elif element[0].lower() == 'cc':
                                cc = int(element[1])
                            elif element[0].lower() == 'dc':
                                dc = int(element[1])
                            disabled += int(element[4])
                            unallocate += int(element[9])
                        nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + 'PDR/CC/DC UP status %d%%/%d%%/%d%%' % (pdr, cc, dc)
                        if disabled > 1 or unallocate > 1:
                            nextStr.Severity = Severity.Critical
                        elif disabled > 0 or unallocate > 0 and nextStr.Severity.value[0] > Severity.Major.value[0]:
                            nextStr.Severity = Severity.Major
                if check[Check.Command.value] == self.checks[7][Check.Command.value]:
                    if elementRE.search(outputLines):
                        saaa, sabb, saper, ucaaa, ucabb, ucaper = 0, 0, 0.0, 0, 0, 0.0
                        for element in elementRE.findall(outputLines):
                            if element[0].lower().strip(' ') == 'site availability':
                                saaa = int(element[1])
                                sabb = int(element[2])
                                saper = float(element[3])
                            elif element[0].lower().strip(' ') == 'unlocked cell availability':
                                ucaaa = int(element[1])
                                ucabb = int(element[2])
                                ucaper = float(element[3])
                        if sabb - saaa >= 5 and ucaper >= 0 and ucaper <= 90:
                            if ucabb - ucaaa >= 40:
                                nextStr.Severity = Severity.Critical
                            elif ucabb - ucaaa >= 20 and nextStr.Severity.value[0] > Severity.Major.value[0]:
                                nextStr.Severity = Severity.Major
                            elif ucabb - ucaaa >= 10 and nextStr.Severity.value[0] > Severity.Minor.value[0]:
                                nextStr.Severity = Severity.Minor
                            elif nextStr.Severity.value[0] > Severity.Warning.value[0]:
                                nextStr.Severity = Severity.Warning
                        nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + ('%d of %d sites are' +
                                                ' fully operational (%3.2f %%)\n%d of %d unlocked cells are up (%3.2f %%)') %\
                                                (saaa, sabb, saper, ucaaa, ucabb, ucaper)
                if check[Check.Command.value] == self.checks[8][Check.Command.value]:
                    if elementRE.search(outputLines):
                        for element in elementRE.findall(outputLines):
                            if (element[0].lower().find('sccpaplocal=ranaplocal') >= 0 or
                                re.search(r'(?i)CnOperator=.*, (IuLink=1,Ranap=.*CS|IuLink=2,Ranap=.*PS)',
                                          element[0].lower()) is not None):
                                nextStr.Severity = Severity.Critical
                    nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + 'RANAP is OK'
                if check[Check.Command.value] == self.checks[9][Check.Command.value]:
                    if elementRE.search(outputLines):
                        for element in elementRE.findall(outputLines):
                            if (int(element[0]) >= 30 and int(element[1]) >= 2 and
                                        nextStr.Severity.value[0] > Severity.Major.value[0]):
                                nextStr.Severity = Severity.Major
                            elif ((int(element[0]) >= 30 or int(element[1]) >= 2) and
                                        nextStr.Severity.value[0] > Severity.Minor.value[0]):
                                nextStr.Severity = Severity.Minor
                        nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + "Total: %s CV's, %s UP's" %\
                                                (element[0], element[1])
                if check[Check.Command.value] == self.checks[10][Check.Command.value]:
                    if elementRE.search(outputLines):
                        for element in elementRE.findall(outputLines):
                            if outputLines.lower().find('roamfroeutranetworkdbtable') >= 0:
                                if nextStr.Severity.value[0] > Severity.Major.value[0]:
                                    nextStr.Severity = Severity.Major
                            else:
                                nextStr.Severity = Severity.Critical
                            nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + 'database is NOT OK'
                    else:
                        nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + 'database is OK'
                if check[Check.Command.value] == self.checks[11][Check.Command.value]:
                    if elementRE.search(outputLines):
                        minVer = ''
                        for element in elementRE.findall(outputLines):
                            if element == '14':
                                if nextStr.Severity.value[0] > Severity.Major.value[0]:
                                    nextStr.Severity = Severity.Major
                            elif element >= '13' and element < '14':
                                nextStr.Severity = Severity.Critical
                            minVer = element if minVer == '' or element < minVer else minVer
                        nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + 'Release W%s' % minVer
                if check[Check.Command.value] == self.checks[14][Check.Command.value]:
                    if elementRE.search(outputLines):
                        for element in elementRE.findall(outputLines):
                            pass
                        nextStr.Observation += ('\n' if nextStr.Observation != '' else '') + \
                                               'Total: %d MOs' % len(elementRE.findall(outputLines))
                        if len(elementRE.findall(outputLines)) > 20:
                            if nextStr.Severity.value[0] > Severity.Warning.value[0]:
                                nextStr.Severity = Severity.Warning
                    if nextStr.Observation == '':
                        nextStr.Observation = 'Total: %d MOs' % 0
                if (check[Check.Command.value] == self.checks[15][Check.Command.value] and
                        check[Check.Caption.value] == self.checks[15][Check.Caption.value]):
                    element = elementRE.search(outputLines)
                    if element is None or element.groups()[0].lower() != '0 (ok)':
                        if nextStr.Severity.value[0] > Severity.Minor.value[0]:
                            nextStr.Severity = Severity.Minor
                        nextStr.Observation = ('\n' if nextStr.Observation != '' else '') + 'Health Check is NOK'
                    else:
                        nextStr.Observation = ('\n' if nextStr.Observation != '' else '') + 'Health Check is OK'
                if (check[Check.Command.value] == self.checks[16][Check.Command.value] and
                        check[Check.Caption.value] == self.checks[16][Check.Caption.value]):
                    element = elementRE.search(outputLines)
                    if element is None or element.groups()[0] == '0':
                        nextStr.Severity = Severity.Warning
                        nextStr.Observation = ('\n' if nextStr.Observation != '' else '') + 'Health Check Schedule is NOK'
                    else:
                        nextStr.Observation = ('\n' if nextStr.Observation != '' else '') + 'Health Check Schedule is OK'
            if nextStr.Observation == '':
                nextStr.Observation = 'No alarms'
            print('%s - Done' % nextStr.CheckName)
            self.output.append(nextStr)

    def savexls(self,filename):
        count = 1
        output = os.path.join(self.dirs['outputDir'], filename+'.xlsx')
        while os.path.exists(output):
            output = os.path.join(self.dirs['outputDir'], filename+str(count)+'.xlsx')
            count += 1
        self.wb.save(output)
        return output

    def writexls(self,filename):
        fs_init_row = 6
        self.wb = openpyxl.load_workbook(filename = self.currentTemplate)
        fs = self.wb['Front Sheet']
        for cell in fs._cells.values():
            if cell.comment == 'AutoCopy':
                fs_init_row = cell.row
                cell.comment = None
                break
        file_number = len(os.listdir(self.dirs['inputDir']))
        if file_number > 2:
            fs.copy_rows(fs_init_row, file_number-2, above=False, copy_style=True, fill_formulae=True)
        else:
            print('Is need more than 2 log files!')
            return
        tmpl = self.wb['Controller log template']
        for num, inFile in enumerate(os.listdir(self.dirs['inputDir'])):
            ws = copy.copy(tmpl)
            ws.title = inFile
            self.wb._add_sheet(ws)
        output = self.savexls(filename)
        self.wb = openpyxl.load_workbook(filename=output)
        try:
            fs = self.wb['Front Sheet']
            tmpl = self.wb['Controller log template']
            for num, inFile in enumerate(os.listdir(self.dirs['inputDir'])):
                print(inFile)
                ws = self.wb[inFile]
                with open(os.path.join(self.dirs['inputDir'], inFile), 'r') as f:
                    self.log = f.read()
                self.output = []
                self.parseLog(inFile)
                for cell in ws.rows[0]:
                    cell.value = cell.value.replace('v<#LogDate#>', self.logdate) if cell.value else None
                es = self.wb['Error list. Summary']
                escurrow = 5
                for row in self.output:
                    cur_row = int(row.Order)+5
                    ws.copy_rows(cur_row, 1, above=False, copy_style=True, fill_formulae=True)
                    for cell in ws.rows[cur_row-1]:
                        cell.value = cell.value.replace('v<#CheckName#>', row.CheckName) if cell.value else None
                        cell.value = cell.value.replace('v<#Severity#>', str(row.Severity)) if cell.value else None
                        cell.value = cell.value.replace('v<#Observation#>', row.Observation) if cell.value else None
                        cell.value = cell.value.replace('v<#DateOf#>', row.DateOf) if cell.value else None
                    if row.Severity != Severity.Ok:
                        es.copy_rows(escurrow, 1, above=False, copy_style=True, fill_formulae=True)
                        for cell in es.rows[escurrow-1]:
                            if cell.column == 'A':
                                cell.value = inFile
                            else:
                                cell.value = ws['%s%s' % (get_column_letter(column_index_from_string(cell.column)-1), cur_row)].value
                        escurrow += 1
                # Nulling last row
                for cell in ws.rows[cur_row]:
                    cell.value = ''
                cur_row = num+fs_init_row
                for cell in fs.rows[cur_row-1]:
                    cell.value = cell.value.replace('v<#FileName#>', inFile) if cell.value else None
                    cell.value = cell.value.replace('v<#MaxRow#>', str(ws.max_row)) if cell.value else None
                    formula = re.compile(r'f<#(.*)#>')
                    if formula.search(str(cell.value)):
                        cell.value = formula.sub(r'\1', cell.value).replace(';',',')
                        cell.data_type = Cell.TYPE_FORMULA
            if tmpl: self.wb.remove_sheet(tmpl)
        # except Exception, e:
            # raise e
        # else:
            # pass
        finally:
            self.wb.save(output)


def main():
    zloyB = ZbAnalyser()
    # zloyB.init_alarms()
    # for row in zloyB.alarms:
        # print(row)
    zloyB.writexls('Preemptive_Support_Report_')

    # BSC => GRAN
    # RNC => WRAN
    # zloyB.test('test')

if __name__ == '__main__':
    main()
